import copy
import shutil
from typing import List, Dict, Any
from openpyxl.reader.excel import load_workbook
from configuration import Configuration as Config
import utils
from pathlib import Path
from openpyxl.styles import Font


def append_component_info(component, xlsx_row):
    xlsx_row['Software Title'] = component.name
    xlsx_row['Version Requested'] = component.version
    xlsx_row['Software Publisher'] = utils.get_publisher(component)
    xlsx_row['Github link (if applicable)'] = component.repo_url


def append_generic_info(xlsx_row):
    xlsx_row['Is this an Executable (yes/no)'] = Config.is_executable
    xlsx_row['Are there any dependencies to other software that also must be installed? (yes/no, list dependencies)'] = f"{Config.has_dependencies}, See {Config.sbom_gen_output_file}"
    xlsx_row['Does it create a non-standard file type? If Yes, can it be scanned by security software (i.e., Anti-Virus)? (yes/ no, if yes list file type, if no explain why not)'] = Config.non_standard_file


def append_efoss_info(xlsx_row):
    xlsx_row['eFOSS approval (if no, then cannot be approved, remove from list) (If yes, include link to eFOSS approval page)'] = ""
    xlsx_row['eFOSS link'] = ""


def clone_workbook(src_xlsx: Path, dest_xlsx: Path) -> None:
    if not src_xlsx.exists():
        raise FileNotFoundError(f"Source xlsx not found: {src_xlsx}")
    dest_xlsx.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src_xlsx, dest_xlsx)


def write_headers(ws, headers: List[str]) -> None:
    """Write headers into the first row and make them bold."""
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = Font(bold=True)


def read_headers_from_sheet(ws, header_row: int = 1) -> List[str]:
    """
    Reads header values from row 1 of the existing sheet.
    Filters out blank/None headers.
    """
    raw = [cell.value for cell in ws[header_row]]
    headers: List[str] = []
    for v in raw:
        if v is None:
            continue
        s = str(v).strip()
        if not s:
            continue
        headers.append(s)
    if not headers:
        raise ValueError(f"No headers found in row {header_row} of sheet '{ws.title}'.")
    return headers


def find_last_data_row(ws, start_row: int = 2) -> int:
    """
    Finds the last row that contains any non-empty cell value.
    This helps avoid ws.max_row being inflated by formatting.
    """
    last = ws.max_row
    # Walk upward until we find a row with any actual value
    while last >= start_row:
        row_has_value = False
        for cell in ws[last]:
            v = cell.value
            if v is None:
                continue
            if isinstance(v, str) and not v.strip():
                continue
            row_has_value = True
            break
        if row_has_value:
            return last
        last -= 1
    return 1  # only headers exist (or sheet is effectively empty)


def dict_row_to_excel_row(row_dict: Dict[str, Any], headers: List[str]) -> List[Any]:
    """Align dict values to the sheet header order."""
    return [row_dict.get(h, "") for h in headers]


def build_rows_from_template(template: Dict[str, Any]) -> List[Dict[str, Any]]:
    """
    Creates row dicts by deepcopy-ing the template so default values (e.g., "Hello") carry over,
    then overrides values per item.
    """
    rows: List[Dict[str, Any]] = []

    for component in Config.component_store.get_all_components():
        xlsx_row = copy.deepcopy(template)
        append_component_info(component, xlsx_row)
        append_generic_info(xlsx_row)
        append_efoss_info(xlsx_row)
        rows.append(xlsx_row)

    return rows


def generate_green_sis_xlsx(green_sis_xlsx_path: Path, sheet_name: str = "SW Submissions"):
    try:
        xlsx_row_template = utils.load_json_file(Config.green_sis_row_template)
        wb = load_workbook(green_sis_xlsx_path)

        if sheet_name not in wb.sheetnames:
            raise KeyError(f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")

        ws = wb[sheet_name]
        headers = read_headers_from_sheet(ws, header_row=1)
        rows = build_rows_from_template(xlsx_row_template)
        last_data_row = find_last_data_row(ws, start_row=2)
        write_row_idx = last_data_row + 1

        for row_dict in rows:
            values = dict_row_to_excel_row(row_dict, headers)
            for col_idx, value in enumerate(values, start=1):
                ws.cell(row=write_row_idx, column=col_idx, value=value)
            write_row_idx += 1

        wb.save(green_sis_xlsx_path)

        print(f"Successfully generated file: {green_sis_xlsx_path}")

    except (FileNotFoundError, IOError) as file_err:
        print(f"File error occurred: {file_err}")
    except KeyError as key_err:
        print(f"Key error occurred: {key_err}")
    except Exception as e:
        print(f"An unexpected error occurred: {e}")


def main():
    Config.green_sis_xlsx_file_name = f"{Config.project_name}-{Config.project_version}-green-sis.xlsx"
    green_sis_xlsx_path = Path(Config.root_dir, "output", Config.green_sis_xlsx_file_name)
    clone_workbook(Config.source_green_sis_xlsx_path, green_sis_xlsx_path)
    generate_green_sis_xlsx(green_sis_xlsx_path, Config.source_green_sis_xlsx_sheet_name)



if __name__ == "__main__":
    main()